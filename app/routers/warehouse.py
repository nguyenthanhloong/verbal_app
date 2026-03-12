from fastapi import APIRouter, Depends, HTTPException
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, func
from datetime import date, timedelta, datetime
from ..database import get_db
from ..models import User, ImportNew, ExportNew, ImportOld, ExportOld, ImportThuong, ExportThuong, ImportLe, ExportLe
from ..schemas import ImportNewCreate, ExportNewCreate, ImportOldCreate, ExportOldCreate, ImportThuongCreate, ExportThuongCreate, ImportLeCreate, ExportLeCreate

from ..dependencies import get_current_user, RequirePermission 

router = APIRouter()

# ==============================================================================
# HÀM BỔ TRỢ (HELPERS)
# ==============================================================================
def get_or_create_device_id(db: Session, serial: str) -> int:
    existing_record = db.query(ImportNew).filter(ImportNew.serial_moi == serial).first()
    if existing_record:
        return existing_record.id
    max_id = db.query(func.max(ImportNew.id)).scalar() or 0
    return max_id + 1

def get_thuong_product_id(db: Session, ma_san_pham: str) -> int:
    existing = db.query(ImportThuong).filter(ImportThuong.ma_san_pham == ma_san_pham).first()
    if existing:
        return existing.id 
    max_id = db.query(func.max(ImportThuong.id)).scalar() or 0
    return max_id + 1 

def get_le_product_id(db: Session, ten_san_pham: str) -> int:
    existing = db.query(ImportLe).filter(ImportLe.ten_san_pham == ten_san_pham).first()
    if existing:
        return existing.id
    max_id = db.query(func.max(ImportLe.id)).scalar() or 0
    return max_id + 1

def is_admin(user: User) -> bool:
    return any(perm.code == "FUNC_ADMIN_ALL" for role in user.roles for perm in role.permissions)

def get_paginated_history(db: Session, model, current_user: User, skip: int, limit: int, options = None):
    """
    Hàm Lịch sử ĐÃ ĐƯỢC MỞ KHÓA thành KHO DÙNG CHUNG.
    Bất kỳ ai có quyền truy cập vào API Lịch sử đều thấy toàn bộ giao dịch của kho đó.
    """
    query = db.query(model)
    
    # Kích hoạt JOIN bảng nếu có truyền options (joinedload)
    if options:
        query = query.options(*options)

    total = query.count()
    records = query.order_by(desc(model.ngay)).offset(skip).limit(limit).all()
    
    # Tự động map dữ liệu thành dạng Flat JSON để VueJS dễ hiển thị
    data = []
    for row in records:
        # Lấy tất cả các cột của bảng hiện tại (ID, số lượng, ghi chú...)
        item = {c.name: getattr(row, c.name) for c in row.__table__.columns}
        
        # Nếu model này có mối quan hệ với Khách hàng, ta bóc tách tên KH ném thẳng vào item
        if hasattr(row, 'customer') and row.customer:
            item['ten_khach_hang'] = row.customer.ten_khach_hang
            item['ma_khach_hang'] = row.customer.ma_khach_hang
        else:
            item['ten_khach_hang'] = None
            item['ma_khach_hang'] = None
            
        data.append(item)
    
    return {"total": total, "data": data}

# ==============================================================================
# PHẦN 1: QUẢN LÝ KHO VIP (QUẢN LÝ THEO TỪNG SERIAL)
# ==============================================================================

@router.post("/api/warehouse/vip/import-new", tags=["Quản lý Kho VIP"], dependencies=[Depends(RequirePermission("FUNC_VIP_NHAP_MOI"))])
def nhap_kho_moi_vip(payload: ImportNewCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    serial = payload.serial_moi
    kiem_tra_serial = db.query(ImportNew).filter(ImportNew.serial_moi == serial).first()
    if kiem_tra_serial:
        raise HTTPException(status_code=400, detail=f"Lỗi: Số Serial '{serial}' này đã tồn tại trong kho! Vui lòng kiểm tra lại.")
        
    device_id = get_or_create_device_id(db, serial)
    new_import = ImportNew(
        id=device_id, serial_moi=serial, ma_kho_spl=payload.ma_kho_spl,
        ma_san_pham=payload.ma_san_pham, ma_may=payload.ma_may, ma_bill=payload.ma_bill,
        nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu,
        pxk_kho_tsb=payload.pxk_kho_tsb, pxk_vp_tsb=payload.pxk_vp_tsb
    )
    db.add(new_import)
    db.commit()
    return {"message": "Nhập kho mới VIP thành công", "id_thiet_bi": device_id}

@router.post("/api/warehouse/vip/export-new", tags=["Quản lý Kho VIP"], dependencies=[Depends(RequirePermission("FUNC_VIP_XUAT_MOI"))])
def xuat_kho_moi_vip(payload: ExportNewCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    serial = payload.serial_moi
    import_record = db.query(ImportNew).filter(ImportNew.serial_moi == serial).first()
    if not import_record:
        raise HTTPException(status_code=400, detail=f"Lỗi: Không tìm thấy Serial '{serial}' trong kho mới!")
        
    kiem_tra_da_xuat = db.query(ExportNew).filter(ExportNew.serial_moi == serial).first()
    if kiem_tra_da_xuat:
        raise HTTPException(status_code=400, detail=f"Lỗi: Serial '{serial}' đã được xuất giao hàng trước đó! Không thể xuất lần 2.")
    
    new_export = ExportNew(
        id=import_record.id, serial_moi=serial, ma_kho_spl=import_record.ma_kho_spl,
        ma_san_pham=import_record.ma_san_pham, ma_may=import_record.ma_may,
        nv_giao_hang=payload.nv_giao_hang, bien_so_xe=payload.bien_so_xe,
        ma_bill=payload.ma_bill, nv_nhap_lieu=current_user.username,
        ghi_chu=payload.ghi_chu, ngay_nhap_kho=import_record.ngay
    )
    db.add(new_export)
    db.commit()
    return {"message": "Xuất kho mới VIP thành công", "id_thiet_bi": import_record.id}

@router.post("/api/warehouse/vip/import-old", tags=["Quản lý Kho VIP"], dependencies=[Depends(RequirePermission("FUNC_VIP_NHAP_CU"))])
def nhap_kho_cu(payload: ImportOldCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    serial_cu = payload.serial_cu
    original_device = db.query(ImportNew).filter(ImportNew.serial_moi == serial_cu).first()
    if not original_device:
        raise HTTPException(status_code=400, detail=f"Lỗi: Không tìm thấy Serial '{serial_cu}'. Hàng này chưa từng được bán ra!")
    
    kiem_tra_da_nhap_cu = db.query(ImportOld).filter(ImportOld.serial_cu == serial_cu).first()
    if kiem_tra_da_nhap_cu:
        raise HTTPException(status_code=400, detail=f"Lỗi: Serial '{serial_cu}' hiện đang nằm trong Kho Trả Hàng rồi! Không thể nhập lại.")
    
    new_import_old = ImportOld(
        id=original_device.id, serial_cu=serial_cu, ma_kho_spl=payload.ma_kho_spl,
        ma_bill=payload.ma_bill, nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu
    )
    db.add(new_import_old)
    db.commit()
    return {"message": "Nhập kho khách trả hàng (Hàng cũ) thành công", "id_thiet_bi": original_device.id}

@router.post("/api/warehouse/vip/export-old", tags=["Quản lý Kho VIP"], dependencies=[Depends(RequirePermission("FUNC_VIP_XUAT_CU"))])
def xuat_kho_cu(payload: ExportOldCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    serial_cu = payload.serial_cu
    import_old_record = db.query(ImportOld).filter(ImportOld.serial_cu == serial_cu).first()
    if not import_old_record:
        raise HTTPException(status_code=400, detail=f"Lỗi: Serial '{serial_cu}' không có mặt trong Kho Trả Hàng!")
        
    kiem_tra_da_xuat_cu = db.query(ExportOld).filter(ExportOld.serial_cu == serial_cu).first()
    if kiem_tra_da_xuat_cu:
        raise HTTPException(status_code=400, detail=f"Lỗi: Serial '{serial_cu}' đã được xuất trả cho Nhà Cung Cấp rồi!")
        
    new_export_old = ExportOld(
        id=import_old_record.id, serial_cu=serial_cu, ma_kho_spl=payload.ma_kho_spl,
        ma_bill=payload.ma_bill, nv_giao_hang=payload.nv_giao_hang, bien_so_xe=payload.bien_so_xe,
        kho_tra_hang=payload.kho_tra_hang, nguoi_nhan=payload.nguoi_nhan,
        nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu, ngay_nhap_kho=import_old_record.ngay
    )
    db.add(new_export_old)
    db.commit()
    return {"message": "Xuất trả nhà cung cấp thành công", "id_thiet_bi": import_old_record.id}

# ==============================================================================
# PHẦN 2: QUẢN LÝ KHO THƯỜNG (QUẢN LÝ THEO LÔ - THUẬT TOÁN FIFO)
# ==============================================================================

@router.post("/api/warehouse/thuong/import", tags=["Quản lý Kho Khách Thường (FIFO)"], dependencies=[Depends(RequirePermission("FUNC_THUONG_NHAP"))])
def nhap_kho_thuong(payload: ImportThuongCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    real_id = get_thuong_product_id(db, payload.ma_san_pham)
    new_batch = ImportThuong(
        id=real_id,customer_id=payload.customer_id, ma_kho_spl=payload.ma_kho_spl, ten_san_pham=payload.ten_san_pham,
        ma_san_pham=payload.ma_san_pham, so_luong=payload.so_luong,
        nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu
    )
    db.add(new_batch)
    db.commit()
    return {"message": f"Đã nhập thành công lô {payload.so_luong} sản phẩm"}

@router.post("/api/warehouse/thuong/export", tags=["Quản lý Kho Khách Thường (FIFO)"], dependencies=[Depends(RequirePermission("FUNC_THUONG_XUAT"))])
def xuat_kho_thuong_fifo(payload: ExportThuongCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    existing_product = db.query(ImportThuong).filter(ImportThuong.ma_san_pham == payload.ma_san_pham).first()
    if not existing_product: raise HTTPException(status_code=400, detail="Sản phẩm này chưa từng được nhập kho!")
    
    real_id = existing_product.id 
    remaining_qty = payload.so_luong
    
    # [FIX BUG]: CHỈ LẤY CÁC LÔ NHẬP TỪ ĐÚNG KHO ĐANG ĐƯỢC YÊU CẦU XUẤT
    import_batches = db.query(ImportThuong)\
        .filter(ImportThuong.id == real_id, ImportThuong.ma_kho_spl == payload.ma_kho_spl)\
        .with_for_update().order_by(ImportThuong.ngay.asc()).all()

    if not import_batches: raise HTTPException(status_code=400, detail=f"Sản phẩm này không có sẵn tại kho {payload.ma_kho_spl}!")
    
    # [FIX BUG]: CHỈ TÍNH TỔNG SỐ ĐÃ XUẤT CỦA SẢN PHẨM NÀY TẠI ĐÚNG KHO ĐÓ
    total_already_exported = db.query(func.coalesce(func.sum(ExportThuong.so_luong), 0))\
        .filter(ExportThuong.id == real_id, ExportThuong.ma_kho_spl == payload.ma_kho_spl).scalar()

    export_records_to_add = []
    cumulative_import = 0

    for batch in import_batches:
        if remaining_qty <= 0: break
        cumulative_import += batch.so_luong
        if total_already_exported >= cumulative_import: continue
            
        stock_available = batch.so_luong
        if total_already_exported > (cumulative_import - batch.so_luong):
            stock_available = cumulative_import - total_already_exported
            
        qty_to_deduct = min(remaining_qty, stock_available)

        new_export = ExportThuong(
            id=real_id,customer_id=payload.customer_id, ma_kho_spl=batch.ma_kho_spl, ten_san_pham=batch.ten_san_pham,
            ma_san_pham=batch.ma_san_pham, so_luong=qty_to_deduct,
            nv_giao_hang=payload.nv_giao_hang, bien_so_xe=payload.bien_so_xe,
            ma_bill=payload.ma_bill, nv_nhap_lieu=current_user.username,
            ghi_chu=payload.ghi_chu, ngay_nhap_kho=batch.ngay 
        )
        export_records_to_add.append(new_export)
        remaining_qty -= qty_to_deduct
        total_already_exported += qty_to_deduct

    # Nếu sau khi vét sạch kho đó mà vẫn chưa đủ hàng -> Báo lỗi
    if remaining_qty > 0:
        raise HTTPException(status_code=400, detail=f"Không đủ hàng tại kho {payload.ma_kho_spl}! Yêu cầu {payload.so_luong} nhưng kho này chỉ còn {payload.so_luong - remaining_qty}.")

    db.add_all(export_records_to_add)
    db.commit()
    return {"message": "Xuất kho FIFO thành công!", "so_luong_yeu_cau": payload.so_luong, "so_lo_da_lay": len(export_records_to_add)}

# ==============================================================================
# PHẦN 3: QUẢN LÝ KHO KHÁCH LẺ (FIFO NHƯNG KHÔNG CÓ MÃ SẢN PHẨM)
# ==============================================================================

@router.post("/api/warehouse/le/import", tags=["Quản lý Kho Khách Lẻ (FIFO)"], dependencies=[Depends(RequirePermission("FUNC_LE_NHAP"))])
def nhap_kho_le(payload: ImportLeCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    real_id = get_le_product_id(db, payload.ten_san_pham) 
    new_batch = ImportLe(
        id=real_id,customer_id=payload.customer_id, ma_kho_spl=payload.ma_kho_spl, ten_san_pham=payload.ten_san_pham,
        so_luong=payload.so_luong, nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu
    )
    db.add(new_batch)
    db.commit()
    return {"message": f"Đã nhập thành công lô {payload.so_luong} sản phẩm khách lẻ"}

@router.post("/api/warehouse/le/export", tags=["Quản lý Kho Khách Lẻ (FIFO)"], dependencies=[Depends(RequirePermission("FUNC_LE_XUAT"))])
def xuat_kho_le_fifo(payload: ExportLeCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    existing_product = db.query(ImportLe).filter(ImportLe.ten_san_pham == payload.ten_san_pham).first()
    if not existing_product: raise HTTPException(status_code=400, detail="Sản phẩm này chưa từng được nhập kho!")
    
    real_id = existing_product.id
    remaining_qty = payload.so_luong
    
    import_batches = db.query(ImportLe).filter(ImportLe.id == real_id)\
        .with_for_update().order_by(ImportLe.ngay.asc()).all()

    if not import_batches: raise HTTPException(status_code=400, detail="Sản phẩm này chưa từng được nhập kho!")

    total_already_exported = db.query(func.coalesce(func.sum(ExportLe.so_luong), 0))\
        .filter(ExportLe.id == real_id).scalar()

    export_records_to_add = []
    cumulative_import = 0

    for batch in import_batches:
        if remaining_qty <= 0: break
        cumulative_import += batch.so_luong
        if total_already_exported >= cumulative_import: continue
            
        stock_available = batch.so_luong
        if total_already_exported > (cumulative_import - batch.so_luong):
            stock_available = cumulative_import - total_already_exported
            
        qty_to_deduct = min(remaining_qty, stock_available)

        new_export = ExportLe(
            id=real_id,customer_id=payload.customer_id, ma_kho_spl=batch.ma_kho_spl, ten_san_pham=batch.ten_san_pham,
            so_luong=qty_to_deduct, nv_giao_hang=payload.nv_giao_hang,
            bien_so_xe=payload.bien_so_xe, ma_bill=payload.ma_bill,
            nv_nhap_lieu=current_user.username, ghi_chu=payload.ghi_chu, ngay_nhap_kho=batch.ngay 
        )
        export_records_to_add.append(new_export)
        remaining_qty -= qty_to_deduct
        total_already_exported += qty_to_deduct

    if remaining_qty > 0:
        raise HTTPException(status_code=400, detail=f"Không đủ hàng trong kho lẻ! Cần {payload.so_luong} nhưng kho chỉ còn {payload.so_luong - remaining_qty}.")

    db.add_all(export_records_to_add)
    db.commit()
    return {"message": "Xuất kho FIFO cho khách lẻ thành công!", "so_luong_yeu_cau": payload.so_luong, "so_lo_da_lay": len(export_records_to_add)}

# ==============================================================================
# PHẦN 4: API LẤY LỊCH SỬ GIAO DỊCH
# ==============================================================================

@router.get("/api/warehouse/vip/import-new", tags=["Lịch Sử VIP"])
def history_vip_import_new(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ImportNew, current_user, skip, limit)

@router.get("/api/warehouse/vip/export-new", tags=["Lịch Sử VIP"])
def history_vip_export_new(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ExportNew, current_user, skip, limit)

@router.get("/api/warehouse/vip/import-old", tags=["Lịch Sử VIP"])
def history_vip_import_old(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ImportOld, current_user, skip, limit)

@router.get("/api/warehouse/vip/export-old", tags=["Lịch Sử VIP"])
def history_vip_export_old(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ExportOld, current_user, skip, limit)

@router.get("/api/warehouse/thuong/import", tags=["Lịch Sử Khách Thường"])
def history_thuong_import(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ImportThuong, current_user, skip, limit, options=[joinedload(ImportThuong.customer)])

@router.get("/api/warehouse/thuong/export", tags=["Lịch Sử Khách Thường"])
def history_thuong_export(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ExportThuong, current_user, skip, limit, options=[joinedload(ExportThuong.customer)])

@router.get("/api/warehouse/le/import", tags=["Lịch Sử Khách Lẻ"])
def history_le_import(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ImportLe, current_user, skip, limit, options=[joinedload(ImportLe.customer)])

@router.get("/api/warehouse/le/export", tags=["Lịch Sử Khách Lẻ"])
def history_le_export(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_paginated_history(db, ExportLe, current_user, skip, limit, options=[joinedload(ExportLe.customer)])

# ==============================================================================
# PHẦN 5: API TÍNH TỔNG TỒN KHO HIỆN TẠI (ĐÃ MỞ KHÓA BẢO MẬT & BỎ LỌC USER)
# ==============================================================================

@router.get("/api/warehouse/thuong/ton-kho", tags=["Tồn Kho"])
def get_ton_kho_thuong(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # BẢO MẬT: Bất kỳ ai có quyền NHẬP hoặc XUẤT đều được xem Dashboard Tồn Kho chung
    user_perms = {p.code for role in current_user.roles for p in role.permissions}
    if not is_admin(current_user) and "FUNC_THUONG_NHAP" not in user_perms and "FUNC_THUONG_XUAT" not in user_perms:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xem tồn kho Khách Thường!")
    
    # Gom nhóm Tổng Nhập (Đã xóa bộ lọc cá nhân để hiển thị kho dùng chung)
    import_query = db.query(
        ImportThuong.id, ImportThuong.ma_san_pham, ImportThuong.ten_san_pham,
        func.sum(ImportThuong.so_luong).label('tong_nhap')
    )
    import_subq = import_query.group_by(ImportThuong.id, ImportThuong.ma_san_pham, ImportThuong.ten_san_pham).subquery()

    # Gom nhóm Tổng Xuất
    export_query = db.query(
        ExportThuong.id,
        func.sum(ExportThuong.so_luong).label('tong_xuat')
    )
    export_subq = export_query.group_by(ExportThuong.id).subquery()

    results = db.query(
        import_subq.c.id, import_subq.c.ma_san_pham, import_subq.c.ten_san_pham,
        import_subq.c.tong_nhap, func.coalesce(export_subq.c.tong_xuat, 0).label('tong_xuat')
    ).outerjoin(
        export_subq, import_subq.c.id == export_subq.c.id
    ).all()

    ton_kho_list = []
    for row in results:
        ton_hien_tai = row.tong_nhap - row.tong_xuat
        if ton_hien_tai > 0:
            ton_kho_list.append({
                "id": row.id, "ma_san_pham": row.ma_san_pham, "ten_san_pham": row.ten_san_pham,
                "tong_nhap": int(row.tong_nhap), "tong_xuat": int(row.tong_xuat), "ton_kho": int(ton_hien_tai)
            })
    return {"data": ton_kho_list}


@router.get("/api/warehouse/le/ton-kho", tags=["Tồn Kho"])
def get_ton_kho_le(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    user_perms = {p.code for role in current_user.roles for p in role.permissions}
    if not is_admin(current_user) and "FUNC_LE_NHAP" not in user_perms and "FUNC_LE_XUAT" not in user_perms:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xem tồn kho Khách Lẻ!")
    
    import_query = db.query(
        ImportLe.id, ImportLe.ten_san_pham,
        func.sum(ImportLe.so_luong).label('tong_nhap')
    )
    import_subq = import_query.group_by(ImportLe.id, ImportLe.ten_san_pham).subquery()

    export_query = db.query(
        ExportLe.id,
        func.sum(ExportLe.so_luong).label('tong_xuat')
    )
    export_subq = export_query.group_by(ExportLe.id).subquery()

    results = db.query(
        import_subq.c.id, import_subq.c.ten_san_pham,
        import_subq.c.tong_nhap, func.coalesce(export_subq.c.tong_xuat, 0).label('tong_xuat')
    ).outerjoin(
        export_subq, import_subq.c.id == export_subq.c.id
    ).all()

    ton_kho_list = []
    for row in results:
        ton_hien_tai = row.tong_nhap - row.tong_xuat
        if ton_hien_tai > 0:
            ton_kho_list.append({
                "id": row.id, "ten_san_pham": row.ten_san_pham,
                "tong_nhap": int(row.tong_nhap), "tong_xuat": int(row.tong_xuat), "ton_kho": int(ton_hien_tai)
            })
    return {"data": ton_kho_list}

# ==============================================================================
# TẦNG XUẤT EXCEL (GIỮ NGUYÊN)
# ==============================================================================
# (Bạn giữ nguyên các hàm xuất Excel của bạn ở vị trí này)
@router.get("/api/warehouse/vip/export-old/excel/{ma_bill}", tags=["Xuất Excel"])
def export_excel_vip_old(ma_bill: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    records = db.query(ExportOld, ImportNew.ma_may).outerjoin(ImportNew, ExportOld.serial_cu == ImportNew.serial_moi).filter(ExportOld.ma_bill == ma_bill).all()
    if not records: raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu cho Mã Bill này!")
    first_record = records[0].ExportOld
    total_items = len(records)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XUAT_TRA_HANG_CU"
    font_bold = Font(bold=True)
    font_title = Font(size=16, bold=True)
    font_normal = Font(size=11)
    font_small = Font(size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def safe_str(value): return str(value) if value is not None else ""
    col_widths = [5, 10, 15, 20, 20, 8, 25, 20, 15, 15, 15]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.merge_cells('A1:G1')
    ws['A1'] = "PHIẾU XUẤT KHO TRẢ HÀNG"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Mã bảng kê: {safe_str(first_record.ma_bill)}"
    ws.merge_cells('F2:K2')
    ws['F2'] = f"Nơi phát hành: {safe_str(first_record.ma_kho_spl)}"
    ws.merge_cells('A3:E3')
    ws['A3'] = f"Ngày chuyển hàng: {first_record.ngay.strftime('%d/%m/%Y') if first_record.ngay else ''}"
    ws.merge_cells('F3:K3')
    ws['F3'] = f"Nơi nhận: {safe_str(first_record.kho_tra_hang)}"
    ws.merge_cells('A4:E4')
    ws['A4'] = f"Tài xế, phụ xe: {safe_str(first_record.nv_giao_hang)}"
    ws.merge_cells('F4:K4')
    ws['F4'] = f"Biển số xe: {safe_str(first_record.bien_so_xe)}"
    ws.merge_cells('A5:K5')
    ws['A5'] = f"Ghi chú: {safe_str(first_record.ghi_chu)}"
    ws.merge_cells('A6:E6')
    ws['A6'] = f"Tổng số kiện: {total_items}"
    ws.merge_cells('F6:G6')
    ws['F6'] = f"Tổng trọng lượng thực (kg): {total_items}"
    ws.merge_cells('H6:K6')
    ws['H6'] = f"Tổng trọng lượng quy đổi(kg): {total_items}"
    for row in range(2, 7):
        for col in range(1, 12):
            ws.cell(row=row, column=col).font = font_normal
            ws.cell(row=row, column=col).alignment = align_left
    headers = ["STT", "ID", "Mã máy cũ", "SỐ SERIAL CŨ", "MÃ BILL", "S.kiện", "Địa chỉ kho trả hàng", "Người nhận hàng", "Ký nhận", "Ghi chú", "Ngày"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header_title)
        cell.font = font_bold
        cell.alignment = align_center
    start_row = 8
    for idx, (exp, ma_may) in enumerate(records, 1):
        row_data = [idx, safe_str(exp.id), safe_str(ma_may), safe_str(exp.serial_cu), safe_str(exp.ma_bill), 1, safe_str(exp.kho_tra_hang), safe_str(exp.nguoi_nhan), "", safe_str(exp.ghi_chu), exp.ngay.strftime('%d/%m/%Y') if exp.ngay else ""]
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num, value=val)
            if col_num in [1, 2, 6, 9, 11]: cell.alignment = align_center
            else: cell.alignment = align_left
        start_row += 1
    sig_row = start_row
    sig_blocks = [("A", "C", "Nhân Viên lập bảng kê"), ("D", "F", "Nhân viên giao hàng"), ("G", "H", "Phụ xe"), ("I", "K", "Người nhận hàng")]
    for start_col, end_col, title in sig_blocks:
        ws.merge_cells(f'{start_col}{sig_row}:{end_col}{sig_row}')
        cell1 = ws[f'{start_col}{sig_row}']
        cell1.value = title
        cell1.font = font_bold
        cell1.alignment = align_center
        ws.merge_cells(f'{start_col}{sig_row+1}:{end_col}{sig_row+1}')
        cell2 = ws[f'{start_col}{sig_row+1}']
        cell2.value = "(Ký và ghi rõ họ tên)"
        cell2.font = font_small
        cell2.alignment = align_center
    ws.row_dimensions[sig_row+1].height = 25
    max_row = sig_row + 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=11):
        for cell in row: cell.border = border_thin
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Phieu_Xuat_Tra_{ma_bill}.xlsx"})

@router.get("/api/warehouse/vip/export-new/excel/{ma_bill}", tags=["Xuất Excel"])
def export_excel_vip_new(ma_bill: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    records = db.query(ExportNew).filter(ExportNew.ma_bill == ma_bill).all()
    if not records: raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu cho Mã Bill này!")
    first_record = records[0]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phieu_Xuat_Giao_Hang"
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def safe_str(value): return str(value) if value is not None else ""
    ws.merge_cells('A1:J1')
    ws['A1'] = "PHIẾU XUẤT GIAO HÀNG"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = align_center
    ws['A3'] = f"Mã bảng kê: {safe_str(first_record.ma_bill)}"
    ws['A4'] = f"Nơi phát hành: {safe_str(first_record.ma_kho_spl)}"
    ws['A5'] = f"Ngày chuyển hàng: {first_record.ngay.strftime('%d/%m/%Y') if first_record.ngay else ''}"
    ws['D3'] = f"Tài xế, phụ xe: {safe_str(first_record.nv_giao_hang)}"
    ws['D4'] = f"Biển số xe: {safe_str(first_record.bien_so_xe)}"
    ws['H3'] = f"Ghi chú: {safe_str(first_record.ghi_chu)}"
    ws['H4'] = f"Tổng số kiện: {len(records)}"
    ws['H5'] = "Tổng trọng lượng thực (kg): "
    ws['H6'] = "Tổng trọng lượng quy đổi (kg): "
    headers = ["STT", "ID", "Mã Sản Phẩm", "Mã Máy", "Số Serial", "MÃ BILL", "S.kiện", "Người nhận hàng", "Ký nhận", "Ghi chú"]
    col_widths = [5, 8, 15, 15, 20, 20, 8, 20, 15, 20]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header_title)
        cell.font = font_bold; cell.alignment = align_center; cell.border = border_thin
    start_row = 9
    for idx, exp in enumerate(records, 1):
        row_data = [idx, safe_str(exp.id), safe_str(exp.ma_san_pham), safe_str(exp.ma_may), safe_str(exp.serial_moi), safe_str(exp.ma_bill), 1, "", "", safe_str(exp.ghi_chu)]
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num, value=val)
            cell.border = border_thin
            if col_num in [1, 2, 7, 9]: cell.alignment = align_center
            else: cell.alignment = align_left
        start_row += 1
    footer_row = start_row + 2
    signatures = [("B", "Nhân Viên lập bảng kê"), ("E", "Nhân viên giao hàng"), ("H", "Người nhận hàng")]
    for col, title in signatures:
        ws[f'{col}{footer_row}'] = title; ws[f'{col}{footer_row}'].font = font_bold; ws[f'{col}{footer_row}'].alignment = align_center
        ws[f'{col}{footer_row+1}'] = "(Ký và ghi rõ họ tên)"; ws[f'{col}{footer_row+1}'].font = Font(italic=True); ws[f'{col}{footer_row+1}'].alignment = align_center
    stream = io.BytesIO()
    wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Phieu_Xuat_Giao_Hang_{ma_bill}.xlsx"})

# TẦNG XUẤT EXCEL KHO THƯỜNG
@router.get("/api/warehouse/thuong/export/excel/{ma_bill}", tags=["Xuất Excel"])
def export_excel_thuong(ma_bill: str, ma_kho_spl: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    clean_bill = ma_bill.strip()
    clean_kho = ma_kho_spl.strip()
    records = db.query(ExportThuong).options(joinedload(ExportThuong.customer)).filter(
        ExportThuong.ma_bill.like(f"%{clean_bill}%"),
        ExportThuong.ma_kho_spl.like(f"%{clean_kho}%")
    ).all()
    
    if not records: raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu!")
    
    first_record = records[0]
    total_qty = sum(r.so_luong for r in records if r.so_luong)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Phieu_Xuat_Kho"
    
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def safe_str(value): return str(value) if value is not None else ""
    ws.merge_cells('A1:H1')
    ws['A1'] = "PHIẾU XUẤT KHO HÀNG HÓA"
    ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = align_center
    ws['A3'] = f"Mã phiếu (Bill): {safe_str(first_record.ma_bill)}"
    ws['A4'] = f"Kho xuất: {safe_str(first_record.ma_kho_spl)}" # Bây giờ nó sẽ in chính xác kho nào ra kho nấy
    ws['A5'] = f"Ngày xuất: {first_record.ngay.strftime('%d/%m/%Y %H:%M') if first_record.ngay else ''}"
    ws['D3'] = f"Tài xế/Giao hàng: {safe_str(first_record.nv_giao_hang)}"
    ws['D4'] = f"Biển số xe: {safe_str(first_record.bien_so_xe)}"
    ws['F3'] = f"Ghi chú: {safe_str(first_record.ghi_chu)}"
    ws['F4'] = f"Tổng số lượng xuất: {total_qty}"
    headers = ["STT", "Mã Lô (ID)", "Khách Hàng", "Mã Sản Phẩm", "Tên Sản Phẩm", "Số Lượng", "ĐVT", "Ghi Chú"]
    col_widths = [5, 12, 30, 20, 35, 12, 10, 25]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header_title)
        cell.font = font_bold; cell.alignment = align_center; cell.border = border_thin
    start_row = 9
    for idx, exp in enumerate(records, 1):
        ten_khach = exp.customer.ten_khach_hang if exp.customer else ""
        row_data = [idx, safe_str(exp.id), ten_khach, safe_str(exp.ma_san_pham), safe_str(exp.ten_san_pham), exp.so_luong, "Kiện/Cái", safe_str(exp.ghi_chu)]
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num, value=val)
            cell.border = border_thin
            cell.alignment = align_center if col_num in [1, 2, 6, 7] else align_left
        start_row += 1
        
    footer_row = start_row + 2
    for col, title in [("B", "Người Lập Phiếu"), ("D", "Tài Xế / Giao Hàng"), ("G", "Người Nhận Hàng")]:
        ws[f'{col}{footer_row}'] = title; ws[f'{col}{footer_row}'].font = font_bold; ws[f'{col}{footer_row}'].alignment = align_center
        ws[f'{col}{footer_row+1}'] = "(Ký và ghi rõ họ tên)"; ws[f'{col}{footer_row+1}'].font = Font(italic=True); ws[f'{col}{footer_row+1}'].alignment = align_center
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Phieu_Xuat_{ma_kho_spl}_{ma_bill}.xlsx"})

# TẦNG XUẤT EXCEL KHO LẺ
@router.get("/api/warehouse/le/export/excel/{ma_bill}", tags=["Xuất Excel"])
def export_excel_le(ma_bill: str, ma_kho_spl: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    clean_bill = ma_bill.strip()
    clean_kho = ma_kho_spl.strip()
    # [FIX BUG]: Thêm điều kiện lọc theo ma_kho_spl
    records = db.query(ExportLe).options(joinedload(ExportLe.customer)).filter(
        ExportLe.ma_bill.like(f"%{clean_bill}%"),
        ExportLe.ma_kho_spl.like(f"%{clean_kho}%")
    ).all()
    
    if not records: raise HTTPException(status_code=404, detail="Không tìm thấy dữ liệu!")
    
    first_record = records[0]
    total_qty = sum(r.so_luong for r in records if r.so_luong)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Phieu_Xuat_Kho"
    
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def safe_str(value): return str(value) if value is not None else ""
    ws.merge_cells('A1:G1')
    ws['A1'] = "PHIẾU XUẤT KHO HÀNG HÓA (LẺ)"
    ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = align_center
    ws['A3'] = f"Mã phiếu (Bill): {safe_str(first_record.ma_bill)}"
    ws['A4'] = f"Kho xuất: {safe_str(first_record.ma_kho_spl)}"
    ws['A5'] = f"Ngày xuất: {first_record.ngay.strftime('%d/%m/%Y %H:%M') if first_record.ngay else ''}"
    ws['D3'] = f"Tài xế/Giao hàng: {safe_str(first_record.nv_giao_hang)}"
    ws['D4'] = f"Biển số xe: {safe_str(first_record.bien_so_xe)}"
    ws['F3'] = f"Tổng số lượng xuất: {total_qty}"
    
    headers = ["STT", "Mã Lô (ID)", "Khách Hàng", "Tên Sản Phẩm", "Số Lượng", "ĐVT", "Ghi Chú"]
    col_widths = [5, 12, 30, 40, 15, 10, 25]
    
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header_title)
        cell.font = font_bold; cell.alignment = align_center; cell.border = border_thin
        
    start_row = 9
    for idx, exp in enumerate(records, 1):
        ten_khach = exp.customer.ten_khach_hang if exp.customer else ""
        row_data = [idx, safe_str(exp.id), ten_khach, safe_str(exp.ten_san_pham), exp.so_luong, "Kiện/Cái", safe_str(exp.ghi_chu)]
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num, value=val)
            cell.border = border_thin
            cell.alignment = align_center if col_num in [1, 2, 5, 6] else align_left
        start_row += 1
        
    footer_row = start_row + 2
    for col, title in [("B", "Người Lập Phiếu"), ("D", "Giao Hàng"), ("G", "Nhận Hàng")]:
        ws[f'{col}{footer_row}'] = title; ws[f'{col}{footer_row}'].font = font_bold; ws[f'{col}{footer_row}'].alignment = align_center
        ws[f'{col}{footer_row+1}'] = "(Ký và ghi rõ họ tên)"; ws[f'{col}{footer_row+1}'].font = Font(italic=True); ws[f'{col}{footer_row+1}'].alignment = align_center
        
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Phieu_Xuat_{ma_kho_spl}_{ma_bill}.xlsx"})

# ==============================================================================
# PHẦN 6: API VẼ BIỂU ĐỒ (CHART DATA - ĐÃ MỞ KHÓA BỘ LỌC CÁ NHÂN)
# ==============================================================================

@router.get("/api/warehouse/thuong/chart", tags=["Thống Kê"])
def get_chart_data_thuong(time_range: str = "7_days", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    today = datetime.now()
    if time_range == "30_days":
        start_date = today - timedelta(days=29)
        num_days = 30
    else:
        start_date = today - timedelta(days=6) 
        num_days = 7
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    # Đã XÓA chốt chặn lọc user để hiển thị toàn bộ hoạt động kho
    import_data = db.query(
        func.date(ImportThuong.ngay).label("ngay_gd"), func.sum(ImportThuong.so_luong).label("tong_nhap")
    ).filter(ImportThuong.ngay >= start_date).group_by(func.date(ImportThuong.ngay)).all()

    export_data = db.query(
        func.date(ExportThuong.ngay).label("ngay_gd"), func.sum(ExportThuong.so_luong).label("tong_xuat")
    ).filter(ExportThuong.ngay >= start_date).group_by(func.date(ExportThuong.ngay)).all()

    import_dict = {str(row.ngay_gd): int(row.tong_nhap) for row in import_data}
    export_dict = {str(row.ngay_gd): int(row.tong_xuat) for row in export_data}

    labels = []
    data_nhap = []
    data_xuat = []
    for i in range(num_days):
        current_step = start_date + timedelta(days=i)
        date_str = current_step.strftime("%Y-%m-%d")
        label_str = current_step.strftime("%d/%m") 
        
        labels.append(label_str)
        data_nhap.append(import_dict.get(date_str, 0))
        data_xuat.append(export_dict.get(date_str, 0))

    return {"labels": labels, "datasets": {"nhap": data_nhap, "xuat": data_xuat}}

@router.get("/api/warehouse/le/chart", tags=["Thống Kê"])
def get_chart_data_le(time_range: str = "7_days", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    today = datetime.now()
    if time_range == "30_days":
        start_date = today - timedelta(days=29)
        num_days = 30
    else:
        start_date = today - timedelta(days=6) 
        num_days = 7
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    # Đã XÓA chốt chặn lọc user
    import_data = db.query(
        func.date(ImportLe.ngay).label("ngay_gd"), func.sum(ImportLe.so_luong).label("tong_nhap")
    ).filter(ImportLe.ngay >= start_date).group_by(func.date(ImportLe.ngay)).all()

    export_data = db.query(
        func.date(ExportLe.ngay).label("ngay_gd"), func.sum(ExportLe.so_luong).label("tong_xuat")
    ).filter(ExportLe.ngay >= start_date).group_by(func.date(ExportLe.ngay)).all()

    import_dict = {str(row.ngay_gd): int(row.tong_nhap) for row in import_data}
    export_dict = {str(row.ngay_gd): int(row.tong_xuat) for row in export_data}

    labels = []
    data_nhap = []
    data_xuat = []
    for i in range(num_days):
        current_step = start_date + timedelta(days=i)
        date_str = current_step.strftime("%Y-%m-%d")
        label_str = current_step.strftime("%d/%m") 
        
        labels.append(label_str)
        data_nhap.append(import_dict.get(date_str, 0))
        data_xuat.append(export_dict.get(date_str, 0))

    return {"labels": labels, "datasets": {"nhap": data_nhap, "xuat": data_xuat}}

@router.get("/api/warehouse/vip/check-serial/{serial}", tags=["Quản lý Kho VIP"])
def check_serial_info(serial: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Tra cứu thông tin thiết bị VIP dựa vào Số Serial.
    Trả về thông tin chi tiết nếu tìm thấy.
    """
    # 1. TÌM TRONG KHO MỚI TRƯỚC (ImportNew)
    device_info = db.query(ImportNew).filter(ImportNew.serial_moi == serial).first()
    
    if not device_info:
        # 2. NẾU KHO MỚI KHÔNG CÓ -> TÌM TIẾP TRONG KHO CŨ (ImportOld)
        device_info_old = db.query(ImportOld).filter(ImportOld.serial_cu == serial).first()
        
        if device_info_old:
            # Vì bảng ImportOld của bạn không lưu Mã SP/Mã Máy, ta phải truy xuất ngược về bảng gốc để lấy
            original_info = db.query(ImportNew).filter(ImportNew.id == device_info_old.id).first()
            if original_info:
                 return {
                    "found": True,
                    "status": "IN_OLD_WAREHOUSE",
                    "data": {
                        "ma_san_pham": original_info.ma_san_pham,
                        "ma_may": original_info.ma_may,
                        "ma_kho_spl": device_info_old.ma_kho_spl,
                        "ma_bill": device_info_old.ma_bill,
                        "pxk_vp": original_info.pxk_vp_tsb,
                        "pxk_kho": original_info.pxk_kho_tsb
                    }
                }
                
        # Nếu quét sạch cả 2 kho mà vẫn không ra
        return {"found": False, "message": "Không tìm thấy số Serial này trong hệ thống."}

    # 3. NẾU TÌM THẤY NGAY Ở KHO MỚI
    return {
        "found": True,
        "status": "IN_NEW_WAREHOUSE",
        "data": {
            "ma_san_pham": device_info.ma_san_pham,
            "ma_may": device_info.ma_may,
            "ma_kho_spl": device_info.ma_kho_spl,
            "ma_bill": device_info.ma_bill,
            "pxk_vp": device_info.pxk_vp_tsb,
            "pxk_kho": device_info.pxk_kho_tsb
        }
    }